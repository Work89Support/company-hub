#!/usr/bin/env python3
"""Prepare a safe Company Hub activity import from the four team workbooks.

The script never connects to Google or Supabase. It reads exported XLSX files,
normalizes Thai/Excel dates and times, reconciles rows against a read-only
identity map exported from Supabase, and writes a JSON payload plus QA summary.
"""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import hashlib
import json
import re
from pathlib import Path

import openpyxl


DOCUMENT_IDS = {
    "01": "1We1ByWq5_tiqVSnVErPNtJcxQxS7Ohw8dw4PDUvARdQ",
    "02": "1wo4hb7opa98uh3QLHHGsXXcZGijaqNHMF80gDjrKhEc",
    "03": "17WaJy8UZLmYHHV_aepQ3fS3_Lko133JLYY6-f9EeXGw",
    "04": "1QhBdwdbUpohnnzMD_zCfgEcyJAj454Ufnn2PFQmzjkg",
}

DEPARTMENTS = {
    "ทีมบริหาร (Management)": ("BOM", "ทีมบริหาร (Management)"),
    "การเงิน (Finance)": ("FIN", "การเงิน (Finance)"),
    "ออดิท (Audit)": ("AUD123", "ออดิท (Audit)"),
    "ทรัพยากรบุคคล (HR)": ("HR", "ทรัพยากรบุคคล (HR)"),
    "ทีม KPI": ("KPI", "ทีม KPI"),
    "เลขานุการ (Secret)": ("SECRET", "เลขานุการ (Secret)"),
    "Content Creative": ("GRAPHIC", "Content Creative"),
    "การตลาด (Marketing)": ("MKT", "การตลาด (Marketing)"),
    "Programmer": ("PROG", "Programmer"),
    "ลูกค้าสัมพันธ์ (CRM)": ("CRM", "ลูกค้าสัมพันธ์ (CRM)"),
    "แอดมิน (Admin) X8": ("ADMIN", "แอดมิน (Admin)"),
    "แอดมิน (Admin) X5": ("ADMIN", "แอดมิน (Admin)"),
    "แอดมิน (Admin) X1": ("ADMIN", "แอดมิน (Admin)"),
    "QC (ตรวจสอบคุณภาพ)": ("QC", "QC (ตรวจสอบคุณภาพ)"),
    "Data Provider": ("BO", "Data Provider"),
}

SPECIAL_SHEETS = {"📘 คู่มือ", "📊 สรุปกลุ่ม", "รายการ"}
UNIT_SEPARATOR = "\x1f"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def raw_text(value: object) -> str:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return clean(value)


def parse_date(value: object) -> tuple[dt.date | None, list[str]]:
    flags: list[str] = []
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        parsed = value
    else:
        parsed = None
        text = clean(value)
        text = re.sub(r"^วันที่\s*", "", text).strip()
        range_match = re.fullmatch(
            r"\d{1,2}/\d{1,2}\s*-\s*(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?",
            text,
        )
        if range_match:
            day = int(range_match.group(1))
            month = int(range_match.group(2))
            year_text = range_match.group(3)
            year = 2026 if not year_text else int(year_text)
            if year < 100:
                year += 2000
            try:
                return dt.date(year, month, day), ["date_range_end"]
            except ValueError:
                return None, ["invalid_date"]
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
            try:
                parsed = dt.datetime.strptime(text, pattern).date()
                break
            except ValueError:
                pass
    if parsed is None:
        return None, ["missing_date" if not clean(value) else "invalid_date"]
    # 69 is commonly entered as the Thai short year 2569. Python parses it as
    # 2069; the other values below are legacy spreadsheet conversion errors.
    if parsed.year in {1969, 2006, 2069, 2569}:
        parsed = parsed.replace(year=2026)
        flags.append("corrected_date")
    if parsed.year != 2026:
        flags.append("unexpected_year")
    return parsed, flags


def parse_time(value: object) -> tuple[str, list[str]]:
    if isinstance(value, dt.datetime):
        value = value.time()
    if isinstance(value, dt.time):
        return value.replace(microsecond=0).isoformat(), []
    if isinstance(value, (int, float)) and 0 <= value < 1:
        seconds = round(float(value) * 86400)
        if 0 <= seconds < 86400:
            return f"{seconds // 3600:02d}:{seconds % 3600 // 60:02d}:{seconds % 60:02d}", ["excel_time"]
    text = clean(value)
    if not text:
        return "", []
    if text.startswith("'"):
        text = text[1:]
    if re.fullmatch(r"\d{1,2}-\d{2}(?:-\d{2})?", text):
        text = text.replace("-", ":")
        separator_flags = ["corrected_separator"]
    elif re.fullmatch(r"\d{1,2};\d{2}(?:;\d{2})?", text):
        text = text.replace(";", ":")
        separator_flags = ["corrected_separator"]
    else:
        separator_flags = []
    match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)(?::([0-5]\d))?", text)
    if match:
        return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}:{int(match.group(3) or 0):02d}", separator_flags
    match = re.fullmatch(r"([01]?\d|2[0-3])\.0", text)
    if match:
        return f"{int(match.group(1)):02d}:00:00", ["thai_decimal_time"]
    match = re.fullmatch(r"([01]?\d|2[0-3])\.([0-5]\d)", text)
    if match:
        return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}:00", ["thai_decimal_time"]
    match = re.fullmatch(r"0?\.\d+", text)
    if match:
        seconds = round(float(text) * 86400)
        if 0 <= seconds < 86400:
            return f"{seconds // 3600:02d}:{seconds % 3600 // 60:02d}:{seconds % 60:02d}", ["excel_time"]
    return "", ["invalid_time"]


def digest(parts: list[str]) -> str:
    return hashlib.md5(UNIT_SEPARATOR.join(parts).encode("utf-8")).hexdigest()


def load_identity_map(path: Path) -> dict[str, collections.deque[str]]:
    result: dict[str, collections.deque[str]] = collections.defaultdict(collections.deque)
    if not path.exists():
        raise FileNotFoundError(f"Existing source-key identity map is required: {path}")
    for line in path.read_text(encoding="utf-8").splitlines():
        identity_hash, source_key = line.split("\t", 1)
        result[identity_hash].append(source_key)
    return result


def group_from_workbook(path: Path, workbook: openpyxl.Workbook) -> str:
    match = re.search(r"กลุ่ม\s+(\d+)", path.name)
    if not match:
        for sheet in workbook.worksheets:
            match = re.search(r"กลุ่ม\s+(\d+)", clean(sheet["A1"].value) + " " + clean(sheet["B2"].value))
            if match:
                break
    if not match:
        raise ValueError(f"cannot determine group from {path}")
    return match.group(1).zfill(2)


def find_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[str, int]]:
    aliases = {
        "วันที่": "date",
        "เว็บที่ดูแล": "site",
        "ชื่อผู้บันทึก": "employee",
        "กิจกรรม / งานที่ทำ": "activity",
        "ประเภทงาน": "category",
        "เวลาเริ่ม": "start",
        "เวลาสิ้นสุด": "end",
        "สถานะ": "status",
        "หมายเหตุ / ผลลัพธ์": "result_note",
        "วันที่ดำเนินการเสร็จ": "completed_date",
        "ปัญหาหน้างาน (ถ้ามี)": "operational_issue",
    }
    for row_number in range(1, min(sheet.max_row, 12) + 1):
        columns: dict[str, int] = {}
        for column in range(1, min(sheet.max_column, 40) + 1):
            label = re.sub(r"\s+", " ", clean(sheet.cell(row_number, column).value))
            if label in aliases:
                columns[aliases[label]] = column
        if {"date", "employee", "activity", "category", "start", "end", "status"}.issubset(columns):
            return row_number, columns
    raise ValueError(f"cannot find activity header in sheet {sheet.title!r}")


def prepare(files: list[Path], identity_map_path: Path) -> tuple[list[dict], dict]:
    identity_map = load_identity_map(identity_map_path)
    original_existing = sum(len(keys) for keys in identity_map.values())
    reserved_keys = {key for keys in identity_map.values() for key in keys}
    payload: list[dict] = []
    rejected: list[dict] = []
    qa = collections.Counter()
    by_date = collections.Counter()
    by_department = collections.Counter()
    matched_source_keys: set[str] = set()

    for path in files:
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
        group = group_from_workbook(path, workbook)
        document_id = DOCUMENT_IDS[group]
        for sheet in workbook.worksheets:
            sheet_name = sheet.title.strip()
            if sheet_name in SPECIAL_SHEETS:
                continue
            if sheet_name not in DEPARTMENTS:
                raise ValueError(f"unknown department sheet: {sheet.title!r}")
            department_code, department_label = DEPARTMENTS[sheet_name]
            header_row, columns = find_columns(sheet)
            last_date: dt.date | None = None
            for row_number in range(header_row + 2, sheet.max_row + 1):
                activity = clean(sheet.cell(row_number, columns["activity"]).value)
                if not activity:
                    continue
                flags: list[str] = []
                raw_date = sheet.cell(row_number, columns["date"]).value
                activity_date, date_flags = parse_date(raw_date)
                if activity_date is not None:
                    last_date = activity_date
                elif not clean(raw_date) and last_date is not None:
                    activity_date = last_date
                    date_flags = ["inherited_date"]
                flags.extend(date_flags)
                employee = clean(sheet.cell(row_number, columns["employee"]).value)
                category = clean(sheet.cell(row_number, columns["category"]).value)
                raw_start = sheet.cell(row_number, columns["start"]).value
                raw_end = sheet.cell(row_number, columns["end"]).value
                start_time, start_flags = parse_time(raw_start)
                end_time, end_flags = parse_time(raw_end)
                status = clean(sheet.cell(row_number, columns["status"]).value)
                result_note = clean(sheet.cell(row_number, columns.get("result_note", 0)).value) if columns.get("result_note") else ""
                completed_date, completed_flags = parse_date(sheet.cell(row_number, columns.get("completed_date", 0)).value) if columns.get("completed_date") else (None, [])
                worksite = clean(sheet.cell(row_number, columns.get("site", 0)).value) if columns.get("site") else ""
                operational_issue = clean(sheet.cell(row_number, columns.get("operational_issue", 0)).value) if columns.get("operational_issue") else ""
                if "missing_date" in completed_flags:
                    completed_date = None
                    completed_flags = []
                flags.extend("start_" + flag for flag in start_flags)
                flags.extend("end_" + flag for flag in end_flags)
                flags.extend("completed_" + flag for flag in completed_flags)
                if not employee:
                    flags.append("missing_employee")
                if not status:
                    flags.append("missing_status")
                if not start_time:
                    flags.append("missing_start_time")
                if not end_time:
                    flags.append("missing_end_time")
                if activity_date is None:
                    rejected.append({"group": group, "sheet": sheet_name, "row": row_number, "reason": "missing_or_invalid_date"})
                    qa["rejected"] += 1
                    continue

                identity_hash = digest([
                    group,
                    department_code,
                    activity_date.isoformat(),
                    employee,
                    activity,
                ])
                sheet_hash = hashlib.sha1(sheet_name.encode("utf-8")).hexdigest()[:10]
                location_key = f"gs-v1-{document_id[:10]}-{sheet_hash}-{row_number}"
                if identity_map.get(identity_hash):
                    candidates = identity_map[identity_hash]
                    if location_key in candidates:
                        source_key = location_key
                        candidates.remove(source_key)
                    else:
                        source_key = candidates.popleft()
                    matched_source_keys.add(source_key)
                    qa["matched_existing"] += 1
                else:
                    source_key = location_key
                    # An edited or shifted sheet row can reuse an occupied
                    # location. Keep its former record and identity intact.
                    if source_key in reserved_keys:
                        source_key = f"{location_key}-{identity_hash}"
                    reserved_keys.add(source_key)
                    qa["new_rows"] += 1

                source_hash = digest([
                    group,
                    department_code,
                    activity_date.isoformat(),
                    employee,
                    activity,
                    category,
                    start_time,
                    end_time,
                    status,
                ])
                record = {
                    "source_key": source_key,
                    "department_code": department_code,
                    "department_label": department_label,
                    "group_code": group,
                    "activity_date": activity_date.isoformat(),
                    "employee_name": employee,
                    "activity": activity,
                    "category": category,
                    "start_time": start_time,
                    "end_time": end_time,
                    "source_start_raw": raw_text(raw_start),
                    "source_end_raw": raw_text(raw_end),
                    "status": status,
                    "source": "Google Sheets Team Sync",
                    "result_note": result_note,
                    "worksite": worksite,
                    "operational_issue": operational_issue,
                    "completed_date": completed_date.isoformat() if completed_date else "",
                    "source_document_id": document_id,
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                    "source_hash": source_hash,
                    "identity_hash": identity_hash,
                    "source_date_raw": clean(raw_date),
                    "data_quality_flags": sorted(set(flags)),
                }
                payload.append(record)
                by_date[activity_date.isoformat()] += 1
                by_department[department_code] += 1
                if flags:
                    qa["rows_with_quality_flags"] += 1
                    for flag in set(flags):
                        qa["flag:" + flag] += 1

    qa["payload_rows"] = len(payload)
    source_key_counts = collections.Counter(row["source_key"] for row in payload)
    duplicate_keys = [key for key, count in source_key_counts.items() if count > 1]
    if duplicate_keys:
        raise ValueError(f"Duplicate source keys require reconciliation: {duplicate_keys[:10]}")
    qa["existing_rows_not_matched"] = original_existing - len(matched_source_keys)
    summary = {
        "totals": dict(sorted(qa.items())),
        "by_date": dict(sorted(by_date.items())),
        "by_department": dict(sorted(by_department.items())),
        "rejected": rejected,
    }
    return payload, summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", action="append", required=True, type=Path)
    parser.add_argument("--identity-map", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--summary", required=True, type=Path)
    args = parser.parse_args()
    payload, summary = prepare(args.input, args.identity_map)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    args.summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
